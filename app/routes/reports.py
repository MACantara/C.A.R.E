from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    jsonify,
    current_app,
    make_response,
)
from flask_login import login_required, current_user
from datetime import datetime, date, timedelta
from app import db
from app.utils.timezone_utils import get_user_timezone, localize_datetime, get_current_time
from app.models.user import User
from app.services.analytics_service import AnalyticsService
from app.utils.sidebar_utils import get_sidebar_stats
from functools import wraps
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")


def admin_or_doctor_required(f):
    """Decorator to require admin or doctor access for reports."""

    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_app.config.get("DISABLE_DATABASE", False):
            flash(
                "Reports are not available in this deployment environment.", "warning"
            )
            return redirect(url_for("main.home"))

        if current_user.role not in ["doctor", "admin"]:
            flash(
                "Access denied. Doctor or admin privileges required for reports.",
                "error",
            )
            return redirect(url_for("main.home"))

        return f(*args, **kwargs)

    return decorated_function


@reports_bp.route("/")
@admin_or_doctor_required
def dashboard():
    """Reports and analytics dashboard."""
    # Get user's timezone for displaying timestamps
    user_timezone = get_user_timezone()
    current_time_local = get_current_time(user_timezone)
    
    # Get sidebar statistics
    stats = get_sidebar_stats()
    
    # Get date range from query params or default to last 30 days
    end_date = date.today()
    start_date = end_date - timedelta(days=30)

    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    if date_from:
        try:
            start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
        except ValueError:
            pass

    if date_to:
        try:
            end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            pass

    # Filter by doctor if not admin
    doctor_filter = None
    if current_user.role == "doctor":
        doctor_filter = current_user.id
    elif request.args.get("doctor_id"):
        doctor_filter = int(request.args.get("doctor_id"))

    # Generate summary metrics
    try:
        appointment_metrics = AnalyticsService.generate_appointment_metrics(
            start_date, end_date, doctor_filter
        )

        prescription_trends = AnalyticsService.generate_prescription_trends(
            start_date, end_date, doctor_filter, limit=5
        )

        diagnosis_trends = AnalyticsService.generate_diagnosis_trends(
            start_date, end_date, doctor_filter, limit=5
        )

        doctor_performance = AnalyticsService.generate_doctor_performance(
            start_date, end_date, doctor_filter
        )

        # Get today's summary
        today_summary = AnalyticsService.generate_daily_summary()

        # Get yesterday's summary for comparison
        yesterday_summary = AnalyticsService.generate_daily_summary(
            date.today() - timedelta(days=1)
        )

    except Exception as e:
        current_app.logger.error(f"Error generating analytics: {e}")
        appointment_metrics = {}
        prescription_trends = []
        diagnosis_trends = []
        doctor_performance = []
        today_summary = {}
        yesterday_summary = {}

    # Get all doctors for filter dropdown (admin only)
    doctors = []
    if current_user.role == "admin":
        doctors = (
            User.query.filter_by(role="doctor", active=True)
            .order_by(User.first_name, User.last_name)
            .all()
        )

    return render_template(
        "medical_dashboard/reports/dashboard.html",
        appointment_metrics=appointment_metrics,
        prescription_trends=prescription_trends,
        diagnosis_trends=diagnosis_trends,
        doctor_performance=doctor_performance,
        today_summary=today_summary,
        yesterday_summary=yesterday_summary,
        doctors=doctors,
        start_date=start_date,
        end_date=end_date,
        doctor_filter=doctor_filter,
        user_timezone=user_timezone,
        current_time_local=current_time_local,
        stats=stats,
    )


@reports_bp.route("/appointments")
@admin_or_doctor_required
def appointment_report():
    """Detailed appointment analytics report."""
    # Get user's timezone for displaying timestamps
    user_timezone = get_user_timezone()
    current_time_local = get_current_time(user_timezone)
    
    # Get sidebar statistics
    stats = get_sidebar_stats()
    
    # Get parameters
    report_type = request.args.get("type", "daily")
    doctor_id = (
        request.args.get("doctor_id")
        if current_user.role == "admin"
        else current_user.id
    )

    # Date range handling
    if report_type == "daily":
        target_date = request.args.get("date")
        if target_date:
            try:
                target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
            except ValueError:
                target_date = date.today()
        else:
            target_date = date.today()

        data = AnalyticsService.generate_daily_summary(target_date)

    elif report_type == "weekly":
        week_start = request.args.get("week_start")
        if week_start:
            try:
                week_start = datetime.strptime(week_start, "%Y-%m-%d").date()
            except ValueError:
                week_start = date.today() - timedelta(days=7)
        else:
            week_start = date.today() - timedelta(days=7)

        data = AnalyticsService.generate_weekly_summary(week_start)

    elif report_type == "monthly":
        year = int(request.args.get("year", date.today().year))
        month = int(request.args.get("month", date.today().month))

        data = AnalyticsService.generate_monthly_summary(year, month)

    # Get all doctors for filter
    doctors = []
    if current_user.role == "admin":
        doctors = User.query.filter_by(role="doctor", active=True).all()

    return render_template(
        "medical_dashboard/reports/appointments.html",
        data=data,
        report_type=report_type,
        doctors=doctors,
        selected_doctor=doctor_id,
        user_timezone=user_timezone,
        current_time_local=current_time_local,
        stats=stats,
    )


@reports_bp.route("/prescriptions")
@admin_or_doctor_required
def prescription_report():
    """Prescription analytics and trends report."""
    # Get user's timezone for displaying timestamps
    user_timezone = get_user_timezone()
    current_time_local = get_current_time(user_timezone)
    
    # Get sidebar statistics
    stats = get_sidebar_stats()
    
    # Get date range
    end_date = date.today()
    start_date = end_date - timedelta(days=30)

    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    if date_from:
        try:
            start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
        except ValueError:
            pass

    if date_to:
        try:
            end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            pass

    doctor_filter = None
    if current_user.role == "doctor":
        doctor_filter = current_user.id
    elif request.args.get("doctor_id"):
        doctor_filter = int(request.args.get("doctor_id"))

    # Generate prescription analytics
    prescription_trends = AnalyticsService.generate_prescription_trends(
        start_date, end_date, doctor_filter, limit=20
    )

    # Get all doctors for filter
    doctors = []
    if current_user.role == "admin":
        doctors = User.query.filter_by(role="doctor", active=True).all()

    return render_template(
        "medical_dashboard/reports/prescriptions.html",
        prescription_trends=prescription_trends,
        doctors=doctors,
        start_date=start_date,
        end_date=end_date,
        doctor_filter=doctor_filter,
        user_timezone=user_timezone,
        current_time_local=current_time_local,
        stats=stats,
    )


@reports_bp.route("/performance")
@admin_or_doctor_required
def performance_report():
    """Doctor performance analytics report."""
    # Get user's timezone for displaying timestamps
    user_timezone = get_user_timezone()
    current_time_local = get_current_time(user_timezone)
    
    # Get sidebar statistics
    stats = get_sidebar_stats()
    
    # Get date range
    end_date = date.today()
    start_date = end_date - timedelta(days=30)

    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    if date_from:
        try:
            start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
        except ValueError:
            pass

    if date_to:
        try:
            end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            pass

    doctor_filter = None
    if current_user.role == "doctor":
        doctor_filter = current_user.id
    elif request.args.get("doctor_id"):
        doctor_filter = int(request.args.get("doctor_id"))

    # Generate performance metrics
    performance_data = AnalyticsService.generate_doctor_performance(
        start_date, end_date, doctor_filter
    )

    # Get all doctors for filter
    doctors = []
    if current_user.role == "admin":
        doctors = User.query.filter_by(role="doctor", active=True).all()

    return render_template(
        "medical_dashboard/reports/performance.html",
        performance_data=performance_data,
        doctors=doctors,
        start_date=start_date,
        end_date=end_date,
        doctor_filter=doctor_filter,
        user_timezone=user_timezone,
        current_time_local=current_time_local,
        stats=stats,
    )


@reports_bp.route("/api/chart-data/<chart_type>")
@admin_or_doctor_required
def api_chart_data(chart_type):
    """API endpoint for chart data."""
    try:
        # Get parameters
        days = int(request.args.get("days", 7))
        doctor_id = (
            request.args.get("doctor_id")
            if current_user.role == "admin"
            else current_user.id
        )

        end_date = date.today()
        start_date = end_date - timedelta(days=days - 1)

        if chart_type == "appointments_timeline":
            # Generate daily appointment counts
            data = []
            current_date = start_date

            while current_date <= end_date:
                daily_summary = AnalyticsService.generate_daily_summary(current_date)
                data.append(
                    {
                        "date": current_date.strftime("%Y-%m-%d"),
                        "appointments": daily_summary["total_appointments"],
                        "completed": daily_summary["completed_appointments"],
                    }
                )
                current_date += timedelta(days=1)

            return jsonify(data)

        elif chart_type == "prescription_trends":
            trends = AnalyticsService.generate_prescription_trends(
                start_date, end_date, doctor_id, limit=10
            )
            return jsonify(trends)

        elif chart_type == "appointment_status":
            metrics = AnalyticsService.generate_appointment_metrics(
                start_date, end_date, doctor_id
            )

            data = [
                {
                    "status": "Completed",
                    "count": metrics.get("completed_appointments", 0),
                },
                {
                    "status": "Scheduled",
                    "count": metrics.get("scheduled_appointments", 0),
                },
                {
                    "status": "Confirmed",
                    "count": metrics.get("confirmed_appointments", 0),
                },
                {
                    "status": "Cancelled",
                    "count": metrics.get("cancelled_appointments", 0),
                },
            ]
            return jsonify(data)

        return jsonify({"error": "Invalid chart type"}), 400

    except Exception as e:
        current_app.logger.error(f"Error generating chart data: {e}")
        return jsonify({"error": "Failed to generate chart data"}), 500


@reports_bp.route("/export/<report_type>")
@admin_or_doctor_required
def export_report(report_type):
    """Export report data in Excel format only."""
    try:
        # Get parameters
        format_type = request.args.get("format", "excel")
        start_date_str = request.args.get("start_date")
        end_date_str = request.args.get("end_date")
        doctor_id = (
            request.args.get("doctor_id")
            if current_user.role == "admin"
            else current_user.id
        )

        # Parse dates
        if start_date_str:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        else:
            start_date = date.today() - timedelta(days=30)

        if end_date_str:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        else:
            end_date = date.today()

        # Only support Excel format
        if format_type != "excel":
            flash("Only Excel export format is supported.", "error")
            return redirect(url_for("reports.dashboard"))

        # Create Excel workbook
        wb = Workbook()

        if report_type == "appointments":
            # Create appointments-focused report
            ws1 = wb.active
            ws1.title = "Appointment Metrics"
            
            # Add appointment metrics
            appointment_metrics = AnalyticsService.generate_appointment_metrics(
                start_date, end_date, doctor_id
            )
            
            # Headers with styling for metrics
            headers = ["Metric", "Value"]
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add appointment metrics data
            metrics_data = [
                ("Total Appointments", appointment_metrics.get('total_appointments', 0)),
                ("Completed Appointments", appointment_metrics.get('completed_appointments', 0)),
                ("Scheduled Appointments", appointment_metrics.get('scheduled_appointments', 0)),
                ("Confirmed Appointments", appointment_metrics.get('confirmed_appointments', 0)),
                ("Cancelled Appointments", appointment_metrics.get('cancelled_appointments', 0)),
                ("Completion Rate (%)", appointment_metrics.get('completion_rate', 0)),
                ("Cancellation Rate (%)", appointment_metrics.get('cancellation_rate', 0)),
                ("Unique Patients", appointment_metrics.get('unique_patients', 0)),
            ]
            
            for row, (metric, value) in enumerate(metrics_data, 2):
                ws1.cell(row=row, column=1, value=metric)
                ws1.cell(row=row, column=2, value=value)

            # Create daily breakdown sheet
            ws2 = wb.create_sheet("Daily Breakdown")
            headers = ["Date", "Total Appointments", "Completed", "Cancelled", "Completion Rate (%)"]
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Generate daily data
            current_date = start_date
            row = 2
            while current_date <= end_date:
                daily_summary = AnalyticsService.generate_daily_summary(current_date)
                ws2.cell(row=row, column=1, value=current_date.strftime("%Y-%m-%d"))
                ws2.cell(row=row, column=2, value=daily_summary.get("total_appointments", 0))
                ws2.cell(row=row, column=3, value=daily_summary.get("completed_appointments", 0))
                ws2.cell(row=row, column=4, value=daily_summary.get("cancelled_appointments", 0))
                completion_rate = 0
                if daily_summary.get("total_appointments", 0) > 0:
                    completion_rate = round((daily_summary.get("completed_appointments", 0) / daily_summary.get("total_appointments", 0)) * 100, 2)
                ws2.cell(row=row, column=5, value=completion_rate)
                current_date += timedelta(days=1)
                row += 1

        elif report_type == "prescriptions":
            # Create prescriptions-focused report
            ws1 = wb.active
            ws1.title = "Prescription Trends"
            
            prescription_trends = AnalyticsService.generate_prescription_trends(
                start_date, end_date, doctor_id, limit=100
            )
            
            # Headers for prescriptions
            headers = ["Medication", "Total Prescriptions", "Unique Patients", "Average per Patient"]
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add prescription data
            for row, med in enumerate(prescription_trends, 2):
                ws1.cell(row=row, column=1, value=med.get('medication', ''))
                ws1.cell(row=row, column=2, value=med.get('total_prescriptions', 0))
                ws1.cell(row=row, column=3, value=med.get('unique_patients', 0))
                avg_per_patient = 0
                if med.get('unique_patients', 0) > 0:
                    avg_per_patient = round(med.get('total_prescriptions', 0) / med.get('unique_patients', 0), 2)
                ws1.cell(row=row, column=4, value=avg_per_patient)

        elif report_type == "performance":
            # Create performance-focused report
            ws1 = wb.active
            ws1.title = "Doctor Performance"
            
            doctor_performance = AnalyticsService.generate_doctor_performance(
                start_date, end_date, doctor_id
            )
            
            # Headers for doctor performance
            headers = ["Doctor Name", "Total Appointments", "Completed", "Cancelled", "Completion Rate (%)", "Cancellation Rate (%)", "Unique Patients"]
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add doctor performance data
            for row, doc in enumerate(doctor_performance, 2):
                ws1.cell(row=row, column=1, value=doc.get('doctor_name', ''))
                ws1.cell(row=row, column=2, value=doc.get('total_appointments', 0))
                ws1.cell(row=row, column=3, value=doc.get('completed_appointments', 0))
                ws1.cell(row=row, column=4, value=doc.get('cancelled_appointments', 0))
                ws1.cell(row=row, column=5, value=doc.get('completion_rate', 0))
                ws1.cell(row=row, column=6, value=doc.get('cancellation_rate', 0))
                ws1.cell(row=row, column=7, value=doc.get('unique_patients', 0))

        elif report_type == "comprehensive":
            # Create comprehensive report with multiple sheets
            # Appointments Summary
            ws1 = wb.active
            ws1.title = "Appointments Summary"
            
            appointment_metrics = AnalyticsService.generate_appointment_metrics(
                start_date, end_date, doctor_id
            )
            
            headers = ["Metric", "Value"]
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            metrics_data = [
                ("Total Appointments", appointment_metrics.get('total_appointments', 0)),
                ("Completed Appointments", appointment_metrics.get('completed_appointments', 0)),
                ("Cancelled Appointments", appointment_metrics.get('cancelled_appointments', 0)),
                ("Completion Rate (%)", appointment_metrics.get('completion_rate', 0)),
                ("Cancellation Rate (%)", appointment_metrics.get('cancellation_rate', 0)),
                ("Unique Patients", appointment_metrics.get('unique_patients', 0)),
            ]
            
            for row, (metric, value) in enumerate(metrics_data, 2):
                ws1.cell(row=row, column=1, value=metric)
                ws1.cell(row=row, column=2, value=value)
            
            # Prescriptions Sheet
            ws2 = wb.create_sheet("Top Prescriptions")
            prescription_trends = AnalyticsService.generate_prescription_trends(
                start_date, end_date, doctor_id, limit=50
            )
            
            headers = ["Medication", "Total Prescriptions", "Unique Patients"]
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, med in enumerate(prescription_trends, 2):
                ws2.cell(row=row, column=1, value=med.get('medication', ''))
                ws2.cell(row=row, column=2, value=med.get('total_prescriptions', 0))
                ws2.cell(row=row, column=3, value=med.get('unique_patients', 0))
            
            # Doctor Performance Sheet
            ws3 = wb.create_sheet("Doctor Performance")
            doctor_performance = AnalyticsService.generate_doctor_performance(
                start_date, end_date, doctor_id
            )
            
            headers = ["Doctor Name", "Total Appointments", "Completed", "Completion Rate (%)", "Unique Patients"]
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, doc in enumerate(doctor_performance, 2):
                ws3.cell(row=row, column=1, value=doc.get('doctor_name', ''))
                ws3.cell(row=row, column=2, value=doc.get('total_appointments', 0))
                ws3.cell(row=row, column=3, value=doc.get('completed_appointments', 0))
                ws3.cell(row=row, column=4, value=doc.get('completion_rate', 0))
                ws3.cell(row=row, column=5, value=doc.get('unique_patients', 0))

        else:
            flash(f"Unknown report type: {report_type}", "error")
            return redirect(url_for("reports.dashboard"))
        
        # Auto-adjust column widths for all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response.headers["Content-Disposition"] = (
            f"attachment; filename={report_type}_report_{start_date}_{end_date}.xlsx"
        )
        return response

    except Exception as e:
        current_app.logger.error(f"Error exporting report: {e}")
        flash("Error exporting report. Please try again.", "error")
        return redirect(url_for("reports.dashboard"))
